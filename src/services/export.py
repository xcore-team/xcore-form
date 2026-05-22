"""
Export des soumissions XForm — Excel, CSV, JSON.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any, Dict, List

from ..schemas.form import FormDefinition, FormSubmission

logger = logging.getLogger("xform.export")


class XFormExporter:

    def export_xlsx(
        self, form: FormDefinition, submissions: List[FormSubmission]
    ) -> bytes:
        """Génère un fichier Excel avec une ligne par soumission."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl requis — pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Soumissions"

        # En-têtes
        headers = ["ID", "Date soumission", "Statut"]
        for field in form.fields:
            headers.append(field.label)
        headers += ["IP", "Durée (s)"]

        header_fill = PatternFill("solid", fgColor="3B82F6")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = max(len(header) + 4, 15)

        # Données
        for row_idx, sub in enumerate(submissions, 2):
            row = [
                sub.id,
                sub.created_at.strftime("%Y-%m-%d %H:%M") if sub.created_at else "",
                sub.status.value,
            ]
            for field in form.fields:
                value = sub.data.get(field.name) or sub.data.get(field.id, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                row.append(str(value) if value is not None else "")

            row += [
                sub.meta.ip or "",
                sub.meta.duration_sec or "",
            ]

            for col, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=val)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_csv(
        self, form: FormDefinition, submissions: List[FormSubmission]
    ) -> str:
        """Génère un CSV UTF-8."""
        output = io.StringIO()
        headers = ["id", "created_at", "status"]
        for field in form.fields:
            headers.append(field.name)
        headers += ["ip", "duration_sec"]

        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()

        for sub in submissions:
            row: Dict[str, Any] = {
                "id": sub.id,
                "created_at": sub.created_at.isoformat() if sub.created_at else "",
                "status": sub.status.value,
                "ip": sub.meta.ip or "",
                "duration_sec": sub.meta.duration_sec or "",
            }
            for field in form.fields:
                value = sub.data.get(field.name) or sub.data.get(field.id, "")
                if isinstance(value, list):
                    value = "|".join(str(v) for v in value)
                row[field.name] = value
            writer.writerow(row)

        return output.getvalue()

    def export_json(
        self, form: FormDefinition, submissions: List[FormSubmission]
    ) -> str:
        """Génère un JSON structuré."""
        data = {
            "form": {
                "id": form.id,
                "title": form.title,
                "slug": form.slug,
                "exported_at": datetime.utcnow().isoformat(),
            },
            "total": len(submissions),
            "submissions": [
                {
                    "id": s.id,
                    "created_at": s.created_at.isoformat() if s.created_at else None,
                    "status": s.status.value,
                    "data": s.data,
                    "meta": {
                        "ip": s.meta.ip,
                        "duration_sec": s.meta.duration_sec,
                    },
                }
                for s in submissions
            ],
        }
        return json.dumps(data, ensure_ascii=False, indent=2)